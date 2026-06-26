"""Jurisdiction threshold screening endpoints."""

from __future__ import annotations

import os
from typing import Any, Optional

from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel

from app.shared.core.config import settings
from app.screening.models.jurisdiction import JurisdictionRule
from app.screening.services.jurisdiction_data_service import load_bundle, verification_metadata
from app.screening.services.threshold_engine import (
    DealParameters,
    RevenueByScope,
    JurisdictionScreeningResult,
    load_all_jurisdictions,
    load_jurisdiction,
    screen_jurisdiction,
)

router = APIRouter(prefix="/jurisdictions", tags=["jurisdictions"])

DATA_DIR = os.path.join(os.path.dirname(settings.data_cases_path), "jurisdictions")


def _get_rules() -> list[JurisdictionRule]:
    try:
        return load_all_jurisdictions(DATA_DIR)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load jurisdiction data: {e}")


# ---------------------------------------------------------------------------
# Request / response models
# ---------------------------------------------------------------------------

class RevenueByScopeInput(BaseModel):
    worldwide: Optional[float] = None
    domestic: Optional[float] = None
    eu_eea: Optional[float] = None
    uk: Optional[float] = None
    us: Optional[float] = None
    by_country: dict[str, float] = {}


class ScreeningRequest(BaseModel):
    acquirer: RevenueByScopeInput = RevenueByScopeInput()
    target: RevenueByScopeInput = RevenueByScopeInput()
    acquirer_assets: Optional[float] = None
    target_assets: Optional[float] = None
    # Per-country assets keyed by jurisdiction_id (e.g. {"ca": 500_000_000})
    acquirer_assets_by_country: dict[str, float] = {}
    target_assets_by_country: dict[str, float] = {}
    deal_value: Optional[float] = None
    deal_currency: str = "EUR"
    revenue_currency: str = "EUR"
    fx_rates: dict[str, float] = {}
    # Transaction structure — used for scope pre-filtering
    deal_type: Optional[str] = None          # merger | share_acquisition | asset_acquisition | joint_venture | minority_stake
    pct_shares_acquired: Optional[float] = None   # 0–100
    post_closing_control: Optional[str] = None    # sole_control | joint_control | material_influence | no_control
    relationship_type: Optional[str] = None  # horizontal | vertical | conglomerate
    combined_market_share: dict[str, float] = {}
    acquirer_market_share: dict[str, float] = {}
    incremental_share: dict[str, float] = {}


class ConditionResultResponse(BaseModel):
    condition_id: str
    met: Optional[bool]
    actual_value: Optional[float]
    threshold_value: float
    gap: Optional[float]
    note: Optional[str] = None
    missing_data: Optional[str] = None


class TestResultResponse(BaseModel):
    test_id: str
    fired: Optional[bool]
    description: Optional[str] = None
    excluded: bool
    exclusion_reason: Optional[str]
    conditions: list[ConditionResultResponse]


class LegalCitationResponse(BaseModel):
    citation: str
    url: Optional[str] = None


class ScreeningResultResponse(BaseModel):
    jurisdiction_id: str
    jurisdiction_name: str
    status: str
    triggered_by: list[str]
    confidence: str
    screening_confidence: str
    source_verification_tier: int = 0
    regression_status: str = "not_run"
    freshness_status: str = "unknown"
    filing_type: Optional[str]
    suspensory: Optional[bool]
    test_results: list[TestResultResponse]
    notes: list[str]
    legal_basis: list[LegalCitationResponse] = []
    authority_url: Optional[str] = None


def _to_deal(req: ScreeningRequest) -> DealParameters:
    def _rev(r: RevenueByScopeInput) -> RevenueByScope:
        return RevenueByScope(
            worldwide=r.worldwide,
            domestic=r.domestic,
            eu_eea=r.eu_eea,
            uk=r.uk,
            us=r.us,
            by_country=r.by_country,
        )

    return DealParameters(
        acquirer=_rev(req.acquirer),
        target=_rev(req.target),
        acquirer_assets=req.acquirer_assets,
        target_assets=req.target_assets,
        deal_value=req.deal_value,
        deal_currency=req.deal_currency,
        deal_type=req.deal_type,
        pct_shares_acquired=req.pct_shares_acquired,
        post_closing_control=req.post_closing_control,
        relationship_type=req.relationship_type,
        revenue_currency=req.revenue_currency,
        fx_rates=req.fx_rates,
        combined_market_share=req.combined_market_share,
        acquirer_market_share=req.acquirer_market_share,
        incremental_share=req.incremental_share,
    )


def _to_deal_for_jurisdiction(req: ScreeningRequest, jurisdiction_id: str) -> DealParameters:
    """Build a DealParameters scoped to a specific jurisdiction.

    For jurisdictions using scope=domestic thresholds, populate the domestic field
    from by_country[jurisdiction_id] when the caller didn't set domestic explicitly.
    Per-country assets override the global acquirer_assets / target_assets.
    """
    jid = jurisdiction_id.lower()

    def _rev(r: RevenueByScopeInput) -> RevenueByScope:
        domestic = r.domestic
        if domestic is None:
            # Look up country-specific revenue as the domestic value for this jurisdiction.
            # Check lowercase first (how the frontend sends it), then uppercase.
            v = r.by_country.get(jid)
            if v is None:
                v = r.by_country.get(jid.upper())
            domestic = v  # stays None if not provided — engine will flag as missing
        return RevenueByScope(
            worldwide=r.worldwide,
            domestic=domestic,
            eu_eea=r.eu_eea,
            uk=r.uk,
            us=r.us,
            by_country=r.by_country,
        )

    acq_assets = req.acquirer_assets_by_country.get(jid, req.acquirer_assets)
    tgt_assets = req.target_assets_by_country.get(jid, req.target_assets)

    return DealParameters(
        acquirer=_rev(req.acquirer),
        target=_rev(req.target),
        acquirer_assets=acq_assets,
        target_assets=tgt_assets,
        deal_value=req.deal_value,
        deal_currency=req.deal_currency,
        deal_type=req.deal_type,
        pct_shares_acquired=req.pct_shares_acquired,
        post_closing_control=req.post_closing_control,
        relationship_type=req.relationship_type,
        revenue_currency=req.revenue_currency,
        fx_rates=req.fx_rates,
        combined_market_share=req.combined_market_share,
        acquirer_market_share=req.acquirer_market_share,
        incremental_share=req.incremental_share,
    )


def _serialise(r: JurisdictionScreeningResult) -> ScreeningResultResponse:
    bundle = load_bundle(DATA_DIR, r.jurisdiction_id)
    meta = verification_metadata(bundle)
    return ScreeningResultResponse(
        jurisdiction_id=r.jurisdiction_id,
        jurisdiction_name=r.jurisdiction_name,
        status=r.status.value,
        triggered_by=r.triggered_by,
        confidence=r.confidence,
        screening_confidence=r.confidence,
        source_verification_tier=meta["source_verification_tier"],
        regression_status=meta["regression_status"],
        freshness_status=meta["freshness_status"],
        filing_type=r.filing_type,
        suspensory=r.suspensory,
        test_results=[
            TestResultResponse(
                test_id=t.test_id,
                fired=t.fired,
                description=t.description,
                excluded=t.excluded,
                exclusion_reason=t.exclusion_reason,
                conditions=[
                    ConditionResultResponse(
                        condition_id=c.condition_id,
                        met=c.met,
                        actual_value=c.actual_value,
                        threshold_value=c.threshold_value,
                        gap=c.gap,
                        note=c.note,
                        missing_data=c.missing_data,
                    )
                    for c in t.conditions
                ],
            )
            for t in r.test_results
        ],
        notes=r.notes,
        legal_basis=[
            LegalCitationResponse(citation=lb.citation, url=lb.url)
            for lb in r.legal_basis
        ],
        authority_url=r.authority_url,
    )


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("/", response_model=list[dict[str, Any]])
def list_jurisdictions():
    """List all available jurisdiction rules (metadata only)."""
    rules = _get_rules()
    return [
        {
            "jurisdiction_id": r.jurisdiction_id,
            "jurisdiction_name": r.jurisdiction_name,
            "authority": r.authority.abbreviation,
            "mandatory": r.regime.mandatory,
            "suspensory": r.regime.suspensory,
            "last_verified": r.last_verified.isoformat(),
            "test_count": len(r.threshold_tests),
        }
        for r in rules
    ]


@router.get("/{jurisdiction_id}", response_model=dict[str, Any])
def get_jurisdiction(jurisdiction_id: str):
    """Return the full rule set for a single jurisdiction."""
    try:
        bundle = load_bundle(DATA_DIR, jurisdiction_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Jurisdiction '{jurisdiction_id}' not found")
    payload = bundle.rule.model_dump()
    payload["verification"] = verification_metadata(bundle)
    return payload


@router.get("/{jurisdiction_id}/passages", response_model=list[dict[str, Any]])
def get_jurisdiction_passages(jurisdiction_id: str):
    """Return source passages (verbatim statutory quotes) for a jurisdiction."""
    try:
        rule = load_jurisdiction(jurisdiction_id, DATA_DIR)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Jurisdiction '{jurisdiction_id}' not found")
    return [p.model_dump() for p in rule.source_passages]


@router.post("/screen", response_model=list[ScreeningResultResponse])
def screen_all_jurisdictions(req: ScreeningRequest):
    """Screen a deal against all loaded jurisdictions."""
    rules = _get_rules()
    results = [
        screen_jurisdiction(_to_deal_for_jurisdiction(req, rule.jurisdiction_id), rule)
        for rule in rules
    ]
    return [_serialise(r) for r in results]


@router.post("/screen/{jurisdiction_id}", response_model=ScreeningResultResponse)
def screen_single_jurisdiction(jurisdiction_id: str, req: ScreeningRequest):
    """Screen a deal against a single jurisdiction."""
    try:
        rule = load_jurisdiction(jurisdiction_id, DATA_DIR)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Jurisdiction '{jurisdiction_id}' not found")
    deal = _to_deal_for_jurisdiction(req, jurisdiction_id)
    result = screen_jurisdiction(deal, rule)
    return _serialise(result)


# ---------------------------------------------------------------------------
# Knowledge base chat
# ---------------------------------------------------------------------------

class _KnowledgeChatMessage(BaseModel):
    role: str  # "user" or "assistant"
    content: str

class _KnowledgeChatRequest(BaseModel):
    message: str
    jurisdiction_ids: list[str] = []
    include_cases: bool = False
    history: list[_KnowledgeChatMessage] = []

class _CitationRef(BaseModel):
    n: int
    jurisdiction_id: str
    section_id: str
    label: str

class _KnowledgeChatResponse(BaseModel):
    response: str
    citations: list[_CitationRef] = []

_KNOWLEDGE_SYSTEM = """You are a merger control expert assistant embedded in a competition law database application.

Answer questions about merger notification rules based strictly on the jurisdiction data provided below.

Guidelines:
- Cite specific thresholds precisely: value, currency, metric, party, and geographic scope
- For multi-jurisdiction comparisons, use numbered lists or a structured breakdown
- If the data does not contain the answer, say so clearly — do not invent rules
- Be concise and professional — this is a tool used by competition lawyers
- Every factual claim MUST have an inline [N] citation marker

RESPONSE FORMAT — return ONLY valid JSON, no markdown fences:
{
  "response": "Your answer with inline [1] [2] citation markers on every factual claim.",
  "citations": [
    {"n": 1, "jurisdiction_id": "au", "section_id": "filing-fees", "label": "Australia – Filing fees"}
  ]
}

Available section_ids (use the exact value):
  scope, minority-stakes, threshold-tests, filing, filing-fees,
  review-periods, gun-jumping, fdi-screening, legal-basis, notes

Use the jurisdiction_id exactly as it appears in the data (e.g. "au", "eu", "gb", "us").
Number citations sequentially starting from 1. Reuse the same [N] for the same source."""


def _build_jurisdiction_context(rules: list) -> str:
    import json

    def _fmt(r) -> dict:
        return r.model_dump(
            mode="json",
            exclude={"source_passages", "practitioner_notes"},
        )

    return json.dumps([_fmt(r) for r in rules], indent=2, default=str)


@router.post("/knowledge-chat", response_model=_KnowledgeChatResponse)
def knowledge_chat(req: _KnowledgeChatRequest):
    """Gemini-powered Q&A over the jurisdiction database."""
    import json
    import os
    from google import genai
    from google.genai import types as gtypes

    api_key = settings.google_api_key or os.environ.get("GOOGLE_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=503, detail="GOOGLE_API_KEY not configured")

    rules = _get_rules()
    if req.jurisdiction_ids:
        rules = [r for r in rules if r.jurisdiction_id in req.jurisdiction_ids]
    if not rules:
        raise HTTPException(status_code=400, detail="No matching jurisdictions found")

    context = _build_jurisdiction_context(rules)
    system = f"{_KNOWLEDGE_SYSTEM}\n\n=== JURISDICTION DATA ===\n{context}\n========================"

    history: list[gtypes.Content] = []
    for m in req.history:
        role = "model" if m.role == "assistant" else "user"
        history.append(gtypes.Content(role=role, parts=[gtypes.Part(text=m.content)]))
    history.append(gtypes.Content(role="user", parts=[gtypes.Part(text=req.message)]))

    client = genai.Client(api_key=api_key)

    _MODELS = ["gemini-2.5-flash-preview-05-20", "gemini-2.0-flash", "gemini-flash-latest"]
    response = None
    last_error: Exception | None = None
    for model_name in _MODELS:
        try:
            response = client.models.generate_content(
                model=model_name,
                contents=history,
                config=gtypes.GenerateContentConfig(
                    system_instruction=system,
                    max_output_tokens=2048,
                    temperature=0.2,
                ),
            )
            break
        except Exception as e:
            last_error = e
            continue

    if response is None:
        raise HTTPException(status_code=503, detail=f"Gemini unavailable: {last_error}")

    raw = (response.text or "").strip()

    # Strip markdown fences if present
    if raw.startswith("```"):
        raw = raw.split("```", 2)[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.strip()

    try:
        data = json.loads(raw)
        return _KnowledgeChatResponse(
            response=data.get("response", raw),
            citations=[
                _CitationRef(
                    n=c["n"],
                    jurisdiction_id=c["jurisdiction_id"],
                    section_id=c["section_id"],
                    label=c["label"],
                )
                for c in data.get("citations", [])
                if all(k in c for k in ("n", "jurisdiction_id", "section_id", "label"))
            ],
        )
    except (json.JSONDecodeError, KeyError):
        return _KnowledgeChatResponse(response=raw, citations=[])


# ---------------------------------------------------------------------------
# Chat intake
# ---------------------------------------------------------------------------

_CHAT_SYSTEM_PROMPT = """You are a merger control intake assistant helping competition lawyers screen deals.

Collect the following through natural conversation — ask one topic at a time, be brief, use plain English. Never expose internal field names or enum values to the user. Do NOT ask about country-specific revenues (handled in a later step).

FIELDS TO COLLECT:
1. Deal value and currency (null if unknown or undisclosed)
2. Transaction type
3. Whether the target company is listed on a stock exchange
4. Percentage of shares / voting rights being acquired (only for share or minority deals)
5. Post-closing governance — what control or influence the acquirer will have (classify from their description, do NOT ask them to pick a category)
6. Acquirer total worldwide revenue (last full financial year)
7. Target total worldwide revenue (last full financial year)

TRANSACTION TYPE MAPPING (internal only, never shown to user):
- Full acquisition / merger → "merger"
- Share purchase → "share_acquisition"
- Asset purchase → "asset_acquisition"
- Joint venture → "joint_venture"
- Minority or partial stake → "minority_stake"

POST-CLOSING CONTROL CLASSIFICATION (classify silently from the user's description):
- "sole_control": acquirer will own a majority, or will direct strategy and day-to-day operations alone with no meaningful co-decision requirement
- "joint_control": no party can act unilaterally on strategic decisions (business plan, budget, key hires) — both must agree; or it's a 50/50 JV
- "material_influence": acquirer gets a board seat, the right to appoint a director, veto rights over strategic decisions, or access to non-public management information — but does not control the company
- "no_control": purely financial investment — no board seat, no vetoes over business decisions, no access to commercially sensitive information; rights are limited to standard minority shareholder protections (anti-dilution, tag-along, information rights on annual accounts)

When asking about governance, phrase it naturally, for example:
"What level of influence will you have over [target] after closing? For example, will you have a board seat or veto over major decisions, or is this more of a passive financial investment?"

If the user's description is ambiguous between material_influence and no_control, ask one targeted follow-up: "Will you have a board seat, or the right to veto decisions like the business plan or budget?"

For a merger or full acquisition, infer post_closing_control = "sole_control" without asking.
For a joint venture, infer post_closing_control = "joint_control" without asking.

Set ready: true ONLY when ALL of the following are known:
- acquirer.worldwide_m is non-null
- target.worldwide_m is non-null
- deal_type is non-null
- post_closing_control is non-null
- pct_shares_acquired is non-null IF deal_type is "share_acquisition" or "minority_stake"

ALWAYS respond with valid JSON (no markdown fences):
{
  "message": "your next question or acknowledgement in plain English",
  "extracted": {
    "deal_value_m": <number or null>,
    "deal_currency": "<USD|EUR|GBP|...>",
    "deal_type": "<merger|share_acquisition|asset_acquisition|joint_venture|minority_stake or null>",
    "target_listed": "<listed|unlisted or null>",
    "pct_shares_acquired": <number 0-100 or null>,
    "post_closing_control": "<sole_control|joint_control|material_influence|no_control or null>",
    "is_passive_investment": <true|false|null>,
    "acquirer": { "worldwide_m": <number or null> },
    "target": { "worldwide_m": <number or null> }
  },
  "ready": <true or false>
}

Convert revenues to millions: "€2 billion" → 2000, "$500m" → 500. Use null for unknown.
is_passive_investment: true only if deal_type is "minority_stake" AND post_closing_control is "no_control" AND the user explicitly confirms a passive / financial-only intent."""


class ChatMessage(BaseModel):
    role: str  # "user" or "assistant"
    content: str


class ChatRequest(BaseModel):
    messages: list[ChatMessage]


class ChatResponse(BaseModel):
    message: str
    extracted: dict
    ready: bool


@router.post("/chat", response_model=ChatResponse)
def chat_intake(req: ChatRequest):
    """LLM-powered conversational deal intake. Returns structured parameters when ready."""
    import json
    import os
    from google import genai
    from google.genai import types as gtypes

    api_key = settings.google_api_key or os.environ.get("GOOGLE_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=503, detail="GOOGLE_API_KEY not configured")

    client = genai.Client(api_key=api_key)

    # Build conversation history for Gemini
    # Gemini requires messages to alternate user/model and start with user.
    # For an empty history (first turn), send a stub user message.
    history: list[gtypes.Content] = []
    for m in req.messages:
        role = "model" if m.role == "assistant" else "user"
        history.append(gtypes.Content(role=role, parts=[gtypes.Part(text=m.content)]))

    if not history:
        # Initial call — send a blank user turn so the model opens the conversation
        history = [gtypes.Content(role="user", parts=[gtypes.Part(text="Start")])]

    # Try models in preference order; fall back if rate-limited
    _GEMINI_MODELS = [
        "gemini-flash-lite-latest",
        "gemini-2.0-flash-lite",
        "gemini-2.0-flash",
        "gemini-flash-latest",
    ]
    response = None
    last_error: Exception | None = None
    for model_name in _GEMINI_MODELS:
        try:
            response = client.models.generate_content(
                model=model_name,
                contents=history,
                config=gtypes.GenerateContentConfig(
                    system_instruction=_CHAT_SYSTEM_PROMPT,
                    max_output_tokens=1024,
                    temperature=0.2,
                ),
            )
            break
        except Exception as e:
            last_error = e
            continue
    if response is None:
        raise HTTPException(status_code=503, detail=f"Gemini unavailable: {last_error}")

    raw = response.text.strip() if response.text else ""

    # Strip markdown code fences if present
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.strip()

    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        data = {"message": raw, "extracted": {}, "ready": False}

    return ChatResponse(
        message=data.get("message", ""),
        extracted=data.get("extracted", {}),
        ready=bool(data.get("ready", False)),
    )


# ---------------------------------------------------------------------------
# File-based financial extraction
# ---------------------------------------------------------------------------

_PARSE_FINANCIALS_PROMPT = """You are a financial data extraction assistant. You have been given the text content of a financial document (balance sheet, income statement, spreadsheet, or annual report).

Extract the following figures if present:
- Company name (acquirer or target)
- Total worldwide/global revenue or turnover (most recent full financial year)
- Revenue or turnover broken down by geography/region if available (e.g. EU/Europe, UK, USA/North America, specific countries)
- Total assets
- Currency used

ALWAYS respond with valid JSON in this exact format:
{
  "message": "brief summary of what you found",
  "entities": [
    {
      "name": "<company name or 'Company 1' if unknown>",
      "worldwide_revenue_m": <number in millions or null>,
      "currency": "<USD|EUR|GBP|...>",
      "year": <fiscal year as integer or null>,
      "regional_breakdown": {
        "eu_eea_m": <number or null>,
        "uk_m": <number or null>,
        "us_m": <number or null>,
        "by_country": { "<country_code>": <number>, ... }
      },
      "total_assets_m": <number or null>
    }
  ],
  "notes": "any caveats or uncertainty about the extraction"
}

If multiple companies are present, include one entry per company. Convert all figures to millions in the stated currency. If unsure about a value, use null rather than guessing."""


@router.post("/parse-financials")
async def parse_financials(file: UploadFile = File(...)):
    """Extract financial figures from an uploaded PDF or Excel file using Gemini."""
    import json
    import io
    import os
    from google import genai
    from google.genai import types as gtypes

    api_key = settings.google_api_key or os.environ.get("GOOGLE_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=503, detail="GOOGLE_API_KEY not configured")

    filename = file.filename or ""
    content = await file.read()

    # Extract text from the file
    text_content = ""

    if filename.lower().endswith(".pdf"):
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                pages = []
                for page in pdf.pages[:20]:  # cap at 20 pages
                    t = page.extract_text()
                    if t:
                        pages.append(t)
                text_content = "\n\n".join(pages)
        except Exception as e:
            raise HTTPException(status_code=422, detail=f"Could not parse PDF: {e}")

    elif filename.lower().endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            sheets = []
            for sheet_name in wb.sheetnames[:5]:  # cap at 5 sheets
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(max_row=200, values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    row_str = "\t".join(cells).strip()
                    if row_str:
                        rows.append(row_str)
                if rows:
                    sheets.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))
            text_content = "\n\n".join(sheets)
        except Exception as e:
            raise HTTPException(status_code=422, detail=f"Could not parse Excel: {e}")

    elif filename.lower().endswith(".csv"):
        try:
            text_content = content.decode("utf-8", errors="replace")
        except Exception as e:
            raise HTTPException(status_code=422, detail=f"Could not read CSV: {e}")

    else:
        raise HTTPException(
            status_code=415,
            detail="Unsupported file type. Upload a PDF, Excel (.xlsx/.xls), or CSV file.",
        )

    if not text_content.strip():
        raise HTTPException(status_code=422, detail="No readable text found in the file.")

    # Truncate to ~30k chars to stay within token limits
    if len(text_content) > 30000:
        text_content = text_content[:30000] + "\n\n[... document truncated ...]"

    client = genai.Client(api_key=api_key)

    _GEMINI_MODELS = ["gemini-flash-lite-latest", "gemini-2.0-flash-lite", "gemini-2.0-flash", "gemini-flash-latest"]
    response = None
    last_error: Exception | None = None
    prompt = f"Document filename: {filename}\n\nDocument content:\n{text_content}"

    for model_name in _GEMINI_MODELS:
        try:
            response = client.models.generate_content(
                model=model_name,
                contents=[gtypes.Content(role="user", parts=[gtypes.Part(text=prompt)])],
                config=gtypes.GenerateContentConfig(
                    system_instruction=_PARSE_FINANCIALS_PROMPT,
                    max_output_tokens=2048,
                    temperature=0.1,
                ),
            )
            break
        except Exception as e:
            last_error = e
            continue

    if response is None:
        raise HTTPException(status_code=503, detail=f"Gemini unavailable: {last_error}")

    raw = response.text.strip() if response.text else ""
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.strip()

    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        raise HTTPException(status_code=422, detail=f"Could not parse Gemini response as JSON: {raw[:200]}")

    return data
